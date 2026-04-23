import io
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from database import CostModelEntry, HeadcountEntry, UserListingEntry, OcRawRow


def parse_workbook(file_bytes: bytes, db: Session, update_refs: bool = False) -> dict:
    """
    Load a Cost Management XLSM.

    update_refs=False (default):
        Only replaces OC raw data.  Reference tables (Cost Model, Headcount,
        User Listing) are left untouched so dashboard edits are preserved.
        If any reference table is still empty (first-ever upload), it is
        populated automatically regardless of this flag.

    update_refs=True:
        Re-reads all 4 sheets and replaces everything (use when the Cost Model
        or Headcount sheet itself has changed).
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)

    oc_ws = next((wb[n] for n in wb.sheetnames if 'oc data refresh' in n.lower()), None)
    if oc_ws is None:
        return _parse_management_tab_fallback(wb)

    spread_year = str(oc_ws['J1'].value or '').strip()

    cm_ws = next((wb[n] for n in wb.sheetnames if n.lower() == 'cost model'), None)
    hc_ws = next((wb[n] for n in wb.sheetnames if n.lower() == 'headcount'), None)
    ul_ws = next((wb[n] for n in wb.sheetnames
                  if 'user based listing' in n.lower() or 'user listing' in n.lower()), None)

    refs_empty = db.query(CostModelEntry).first() is None

    if update_refs or refs_empty:
        if cm_ws:
            _load_cost_model(cm_ws, db)
        if hc_ws:
            _load_headcount(hc_ws, db, spread_year)
        if ul_ws:
            _load_user_listing(ul_ws, db)

    _load_oc_data(oc_ws, db)
    wb.close()

    from allocator import run_allocation
    rows = run_allocation(db)
    return {"rows": rows, "rowCount": len(rows), "sheetName": "OC Data Refresh",
            "refs_empty": refs_empty}


# ── Sheet loaders ─────────────────────────────────────────────────────────────

def _load_cost_model(ws, db):
    db.query(CostModelEntry).delete()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 13:
            continue
        branch_name = str(row[0] or '').strip()
        if not branch_name:
            continue
        branch_code = branch_name[:4] if len(branch_name) >= 4 else branch_name
        gl_code = str(row[1] or '').strip()
        pid = str(row[3] or '').strip()
        if not pid:
            continue
        db.add(CostModelEntry(
            branch_name         = branch_name,
            gl_code             = gl_code,
            branch_code         = branch_code,
            pid                 = pid,
            gl_category         = str(row[4] or '').strip(),
            cost_model_category = str(row[5] or '').strip(),
            description         = str(row[6] or '').strip(),
            required            = str(row[7] or '').strip(),
            current_cost_model  = str(row[8] or '').strip(),
            allocation          = str(row[9] or '').strip(),
            future_cost_model   = str(row[10] or '').strip(),
            showback_type       = str(row[11] or '').strip(),
            user_listing_flag   = str(row[12] or '').strip(),
        ))
    db.commit()


def _load_headcount(ws, db, spread_year: str):
    # Row 3 = headers; find column matching spread_year
    hc_col = None
    for c in range(1, 21):
        if str(ws.cell(row=3, column=c).value or '').strip() == spread_year:
            hc_col = c
            break
    # Fallback: rightmost non-empty header after col 3
    if hc_col is None:
        for c in range(20, 3, -1):
            if ws.cell(row=3, column=c).value:
                hc_col = c
                break
    if hc_col is None:
        return

    db.query(HeadcountEntry).delete()
    for r in range(4, 31):
        short_code = str(ws.cell(row=r, column=2).value or '').strip()
        if not short_code or short_code.lower() == 'cir':
            continue
        val = ws.cell(row=r, column=hc_col).value
        try:
            fy2026 = float(val) if val is not None else 0.0
        except (ValueError, TypeError):
            fy2026 = 0.0
        db.add(HeadcountEntry(
            dept_code = str(ws.cell(row=r, column=1).value or '').strip(),
            short_code = short_code,
            dept_name  = str(ws.cell(row=r, column=3).value or '').strip(),
            fy2026     = fy2026,
        ))
    db.commit()


def _load_user_listing(ws, db):
    db.query(UserListingEntry).delete()

    def _f(row, col_0):
        if col_0 < len(row) and row[col_0] is not None:
            try:
                return float(row[col_0])
            except (ValueError, TypeError):
                return 0.0
        return 0.0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 4:
            continue
        branch_name = str(row[0] or '').strip()
        if not branch_name:
            continue
        branch_code = branch_name[:4] if len(branch_name) >= 4 else branch_name
        pid = str(row[3] or '').strip()
        if not pid:
            continue
        # Cols F-R (0-based indices 5-17): CEO,Legal,CorpOps,HR,Audit,CD&O,Finance,Tech,IO,IRR,PE,CM&CI,ISR
        db.add(UserListingEntry(
            branch_name = branch_name,
            branch_code = branch_code,
            gl_code     = str(row[1] or '').strip(),
            pid         = pid,
            description = str(row[4] or '').strip() if len(row) > 4 else '',
            ceo        = _f(row, 5),
            legal      = _f(row, 6),
            corp_ops   = _f(row, 7),
            hr         = _f(row, 8),
            audit      = _f(row, 9),
            cdo        = _f(row, 10),
            finance    = _f(row, 11),
            technology = _f(row, 12),
            io         = _f(row, 13),
            irr        = _f(row, 14),
            pe         = _f(row, 15),
            cmci       = _f(row, 16),
            isr        = _f(row, 17),
        ))
    db.commit()


def _load_oc_data(ws, db):
    db.query(OcRawRow).delete()

    def _n(cell_val):
        if cell_val is None:
            return 0.0
        if isinstance(cell_val, (int, float)):
            return float(cell_val)
        try:
            return float(str(cell_val).replace(',', ''))
        except (ValueError, TypeError):
            return 0.0

    # Auto-detect leaf indent level — scan up to 500 rows and take the maximum
    # indent seen. This handles files where the leaf level is 3 or differs.
    indent_counts: dict[int, int] = {}
    for r in range(4, min(ws.max_row + 1, 500)):
        cell = ws.cell(row=r, column=1)
        if not cell.value:
            continue
        lvl = int((cell.alignment.indent if cell.alignment else 0) or 0)
        if lvl > 0:
            indent_counts[lvl] = indent_counts.get(lvl, 0) + 1
    leaf_indent = max(indent_counts) if indent_counts else 3

    # Column layout (after user added col C to OC Data Refresh):
    #   B(2) = Actuals (current FY)
    #   C(3) = Budget  (current FY)
    #   D(4) = Forecast 1 (next FY)
    #   E(5) = Forecast 2 (FY+2)
    #   F(6) = ignored
    for row_idx in range(4, ws.max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1)
        val = cell_a.value
        if not val:
            continue
        lvl = int((cell_a.alignment.indent if cell_a.alignment else 0) or 0)
        if lvl != leaf_indent:
            continue
        actuals = _n(ws.cell(row=row_idx, column=2).value)
        budget  = _n(ws.cell(row=row_idx, column=3).value)
        f1      = _n(ws.cell(row=row_idx, column=4).value)
        f2      = _n(ws.cell(row=row_idx, column=5).value)
        if actuals == 0 and budget == 0 and f1 == 0 and f2 == 0:
            continue
        db.add(OcRawRow(oc_cell=str(val).strip(),
                        actuals=actuals, budget=budget, forecast1=f1, forecast2=f2))

    db.commit()


# ── Legacy fallback (Management Tab output) ───────────────────────────────────

def _parse_management_tab_fallback(wb) -> dict:
    sheet_name = next((n for n in wb.sheetnames if n.endswith("Management Tab")), wb.sheetnames[0])
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=11, values_only=True):
        if not row or len(row) < 17:
            continue

        def g(col_1):
            idx = col_1 - 1
            return row[idx] if idx < len(row) else None

        def s(col_1):
            v = g(col_1)
            return str(v).strip() if v is not None else ''

        def n(col_1):
            v = g(col_1)
            if v is None:
                return 0.0
            if isinstance(v, (int, float)):
                return float(v)
            try:
                return float(str(v).replace(',', ''))
            except (ValueError, TypeError):
                return 0.0

        actuals, f1, f2, budget = n(14), n(15), n(16), n(17)
        if actuals == 0 and f1 == 0 and f2 == 0 and budget == 0:
            continue
        if not s(2):
            continue

        rows.append({
            'branch': s(2), 'glCode': s(3), 'branchCode': s(4), 'pid': s(5),
            'glCategory': s(6), 'costModelCategory': s(7), 'description': s(8),
            'required': s(9), 'currentCostModel': s(10), 'allocation': s(11),
            'futureCostModel': s(12), 'showbackType': s(13),
            'actuals': actuals, 'forecast1': f1, 'forecast2': f2, 'budget': budget,
            'ceo': n(18), 'legal': n(19), 'hr': n(20), 'audit': n(21),
            'cdo': n(22) * 0.5, 'corpOps': n(22) * 0.5, 'finance': n(23), 'technology': n(24),
            'io': n(25), 'irr': n(26), 'isr': n(27), 'cmci': n(28), 'pe': n(29),
            'comments': s(30),
        })
    wb.close()
    return {'rows': rows, 'rowCount': len(rows), 'sheetName': sheet_name}
