import os, json
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
load_dotenv(Path(__file__).parent.parent / ".env")

from fastapi import FastAPI, Depends, HTTPException, Request, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from starlette.middleware.sessions import SessionMiddleware
from sqlalchemy.orm import Session

from auth import router as auth_router, get_current_user, require_admin
from database import engine, Base, get_db, User, AccessLog, CostModelEntry, HeadcountEntry, UserListingEntry, AppSetting, ShareConfig
from models import UserCreate, UserUpdate, UserResponse, LogEntry
from permissions import filter_for_user
from parser import parse_workbook
from logger import log_access

# Create DB tables on startup
Base.metadata.create_all(bind=engine)

# Migrate: add fy2027/fy2028 to headcount if upgrading from older schema
from sqlalchemy import text as _sql_text
with engine.connect() as _c:
    for _col in ['fy2027', 'fy2028']:
        try:
            _c.execute(_sql_text(f'ALTER TABLE headcount ADD COLUMN {_col} FLOAT DEFAULT 0.0'))
            _c.commit()
        except Exception:
            _c.rollback()
    try:
        _c.execute(_sql_text("ALTER TABLE user_listing ADD COLUMN branch_name TEXT DEFAULT ''"))
        _c.commit()
    except Exception:
        _c.rollback()
    try:
        _c.execute(_sql_text("ALTER TABLE users ADD COLUMN can_edit_user_listing BOOLEAN DEFAULT 0"))
        _c.commit()
    except Exception:
        _c.rollback()
    try:
        _c.execute(_sql_text("ALTER TABLE users ADD COLUMN can_view_quality BOOLEAN DEFAULT 0"))
        _c.commit()
    except Exception:
        _c.rollback()
    try:
        _c.execute(_sql_text("ALTER TABLE users ADD COLUMN allowed_departments JSON DEFAULT '[]'"))
        _c.commit()
    except Exception:
        _c.rollback()
    # share_configs table (create if missing — idempotent via metadata)
    try:
        _c.execute(_sql_text("SELECT id FROM share_configs LIMIT 1"))
    except Exception:
        _c.rollback()
        Base.metadata.tables['share_configs'].create(bind=engine)
        _c.commit()

# ── Production safety checks ──────────────────────────────────────────────────
if os.getenv("APP_ENV") == "production":
    _WEAK = {"", "replace-with-64-char-hex-string",
             "replace-with-a-different-64-char-hex-string",
             "CHANGE_THIS_IN_PRODUCTION"}
    if os.getenv("JWT_SECRET_KEY", "") in _WEAK:
        raise RuntimeError(
            "FATAL: JWT_SECRET_KEY is not set or is still the placeholder. "
            "Generate: python -c \"import secrets; print(secrets.token_hex(32))\""
        )
    if os.getenv("SESSION_SECRET", "") in _WEAK:
        raise RuntimeError(
            "FATAL: SESSION_SECRET is not set or is still the placeholder. "
            "Generate: python -c \"import secrets; print(secrets.token_hex(32))\""
        )
    if os.getenv("STORAGE_BACKEND", "local") == "local":
        import warnings
        warnings.warn(
            "WARNING: STORAGE_BACKEND=local in production — uploaded cost data will be "
            "lost on every container restart. Set STORAGE_BACKEND=azure and configure "
            "AZURE_STORAGE_CONNECTION_STRING.", stacklevel=1,
        )

app = FastAPI(
    title="Technology Showback Dashboard API",
    version="1.0.0",
    docs_url="/api/docs",
    redoc_url="/api/redoc",
)

# ── Scheduled share sends ─────────────────────────────────────────────────────
from apscheduler.schedulers.background import BackgroundScheduler as _BGScheduler

_scheduler = _BGScheduler()


def _check_scheduled_sends():
    from database import SessionLocal
    from reporter import generate_dept_pdf
    from emailer import send_report
    db = SessionLocal()
    try:
        now = datetime.utcnow()
        configs = db.query(ShareConfig).filter(
            ShareConfig.schedule != 'manual',
            ShareConfig.next_run <= now,
            ShareConfig.emails != None,
        ).all()
        for cfg in configs:
            if not cfg.emails:
                continue
            try:
                data = _load_data()
                pdf  = generate_dept_pdf(data.get("rows", []), cfg.department, cfg.period)
                _DLABELS = {'ceo':'CEO','legal':'Legal','hr':'HR','audit':'Audit',
                            'cdo':'CD&O','corpOps':'Corp Ops','finance':'Finance',
                            'technology':'Technology','io':'IO','irr':'IRR',
                            'pe':'PE','cmci':'CM&CI','isr':'ISR'}
                dept_label = _DLABELS.get(cfg.department, cfg.department.title())
                send_report(cfg.emails, dept_label, cfg.period, pdf)
                cfg.last_sent = now
                cfg.next_run  = _next_run(cfg.schedule, now)
                db.commit()
            except Exception as e:
                print(f"[share scheduler] failed for cfg {cfg.id}: {e}")
    finally:
        db.close()


def _next_run(schedule: str, from_dt=None):
    from datetime import timedelta
    base = from_dt or datetime.utcnow()
    if schedule == 'monthly':
        # advance by ~30 days
        return base.replace(day=1) if False else base + timedelta(days=30)
    if schedule == 'quarterly':
        return base + timedelta(days=91)
    return None


@app.on_event("startup")
def _start_scheduler():
    _scheduler.add_job(_check_scheduled_sends, 'interval', hours=1, id='share_scheduler',
                       replace_existing=True)
    _scheduler.start()


@app.on_event("shutdown")
def _stop_scheduler():
    _scheduler.shutdown(wait=False)

# ── Middleware ────────────────────────────────────────────────────────────────
origins = os.getenv("ALLOWED_ORIGINS", "http://localhost:3000,http://localhost:5173").split(",")
app.add_middleware(CORSMiddleware, allow_origins=origins, allow_credentials=True,
                   allow_methods=["*"], allow_headers=["*"])
app.add_middleware(SessionMiddleware,
                   secret_key=os.getenv("SESSION_SECRET", os.urandom(32).hex()))

app.include_router(auth_router, prefix="/auth", tags=["auth"])

# ── Cost data storage ─────────────────────────────────────────────────────────
DATA_FILE = Path(__file__).parent / "data" / "cost_data.json"
DATA_FILE.parent.mkdir(exist_ok=True)
DIST_DIR  = Path(__file__).parent.parent / "dist"

STORAGE_BACKEND = os.getenv("STORAGE_BACKEND", "local")
_AZ_CONN_STR    = os.getenv("AZURE_STORAGE_CONNECTION_STRING", "")
_AZ_CONTAINER   = os.getenv("AZURE_STORAGE_CONTAINER", "showback-data")
_BLOB_NAME      = "cost_data.json"


def _az_client():
    from azure.storage.blob import BlobServiceClient
    svc = BlobServiceClient.from_connection_string(_AZ_CONN_STR)
    c = svc.get_container_client(_AZ_CONTAINER)
    try:
        c.create_container()
    except Exception:
        pass
    return c


def _load_data() -> dict:
    empty = {"rows": [], "updatedAt": None, "sheetName": None}
    if STORAGE_BACKEND == "azure":
        try:
            return json.loads(_az_client().download_blob(_BLOB_NAME).readall())
        except Exception:
            return empty
    if not DATA_FILE.exists():
        return empty
    return json.loads(DATA_FILE.read_text(encoding="utf-8"))


def _migrate_cdo_split():
    """Re-run allocation if saved rows still use the old combined cdoCorpOps field."""
    data = _load_data()
    rows = data.get("rows", [])
    if not rows or 'cdoCorpOps' not in rows[0]:
        return
    from database import SessionLocal
    from allocator import run_allocation_all_periods
    db = SessionLocal()
    try:
        new_rows = run_allocation_all_periods(db)
        if new_rows:
            _save_data(new_rows, data.get("sheetName") or "Cost Data")
    finally:
        db.close()


_migrate_cdo_split()


def _save_data(rows: list, sheet_name: str):
    payload = {"rows": rows, "sheetName": sheet_name, "updatedAt": datetime.utcnow().isoformat()}
    if STORAGE_BACKEND == "azure":
        _az_client().upload_blob(_BLOB_NAME, json.dumps(payload), overwrite=True)
    else:
        DATA_FILE.write_text(json.dumps(payload), encoding="utf-8")
    return payload


# ── Data endpoints ────────────────────────────────────────────────────────────
@app.get("/api/data")
async def get_data(request: Request, current_user: User = Depends(get_current_user),
                   db: Session = Depends(get_db)):
    log_access(db, current_user.email, "view_data", "/api/data",
               request.client.host if request.client else "")
    payload = _load_data()
    filtered = filter_for_user(payload["rows"], current_user)
    return {**payload, "rows": filtered,
            "totalRows": len(payload["rows"]), "visibleRows": len(filtered)}


@app.post("/api/upload")
async def upload_file(request: Request, file: UploadFile = File(...),
                      update_refs: bool = False,
                      current_user: User = Depends(require_admin),
                      db: Session = Depends(get_db)):
    content = await file.read()
    result = parse_workbook(content, db, update_refs=update_refs)
    rows, sheet_name = result["rows"], result["sheetName"]
    _save_data(rows, sheet_name)
    log_access(db, current_user.email, "upload_data", f"{sheet_name} ({len(rows)} rows)",
               request.client.host if request.client else "",
               {"rowCount": len(rows), "filename": file.filename,
                "update_refs": update_refs})
    return {"ok": True, "rowCount": len(rows), "sheetName": sheet_name,
            "refsUpdated": update_refs or result.get("refs_empty", False)}


@app.post("/api/push")
async def push_data(request: Request, current_user: User = Depends(require_admin),
                    db: Session = Depends(get_db)):
    """Finance script POSTs pre-parsed rows here after running the VBA macro."""
    body = await request.json()
    rows       = body.get("rows", [])
    sheet_name = body.get("sheetName", "Unknown")
    if not isinstance(rows, list):
        raise HTTPException(400, "rows must be an array")
    _save_data(rows, sheet_name)
    log_access(db, current_user.email, "push_data", f"{sheet_name} ({len(rows)} rows)",
               request.client.host if request.client else "")
    return {"ok": True, "rowCount": len(rows)}


# ── Admin: user management ────────────────────────────────────────────────────
@app.get("/admin/users", response_model=list[UserResponse])
async def list_users(request: Request, current_user: User = Depends(require_admin),
                     db: Session = Depends(get_db)):
    log_access(db, current_user.email, "list_users", "/admin/users",
               request.client.host if request.client else "")
    return db.query(User).order_by(User.email).all()


@app.post("/admin/users", response_model=UserResponse, status_code=201)
async def create_user(request: Request, user_in: UserCreate,
                      current_user: User = Depends(require_admin),
                      db: Session = Depends(get_db)):
    if db.query(User).filter(User.email == user_in.email.lower()).first():
        raise HTTPException(409, f"{user_in.email} already exists — use PUT to update")
    user = User(
        email=user_in.email.lower(),
        display_name=user_in.display_name or user_in.email.split("@")[0],
        is_admin=user_in.is_admin,
        can_edit_user_listing=user_in.can_edit_user_listing,
        can_view_quality=user_in.can_view_quality,
        allowed_gl_codes=user_in.allowed_gl_codes,
        allowed_branches=user_in.allowed_branches,
        allowed_departments=user_in.allowed_departments,
    )
    db.add(user); db.commit(); db.refresh(user)
    log_access(db, current_user.email, "create_user", user.email,
               request.client.host if request.client else "")
    return user


@app.put("/admin/users/{email}", response_model=UserResponse)
async def update_user(request: Request, email: str, user_in: UserUpdate,
                      current_user: User = Depends(require_admin),
                      db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == email.lower()).first()
    if not user:
        raise HTTPException(404, "User not found")
    for field, val in user_in.model_dump(exclude_unset=True).items():
        setattr(user, field, val)
    user.updated_at = datetime.utcnow()
    db.commit(); db.refresh(user)
    log_access(db, current_user.email, "update_user", email,
               request.client.host if request.client else "")
    return user


@app.delete("/admin/users/{email}")
async def delete_user(request: Request, email: str,
                      current_user: User = Depends(require_admin),
                      db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == email.lower()).first()
    if not user:
        raise HTTPException(404, "User not found")
    db.delete(user); db.commit()
    log_access(db, current_user.email, "delete_user", email,
               request.client.host if request.client else "")
    return {"ok": True}


# ── Admin: usage logs ─────────────────────────────────────────────────────────
@app.get("/admin/logs", response_model=list[LogEntry])
async def get_logs(current_user: User = Depends(require_admin),
                   db: Session = Depends(get_db), limit: int = 500):
    return db.query(AccessLog).order_by(AccessLog.timestamp.desc()).limit(limit).all()


# ── Admin: Cost Model reference table ─────────────────────────────────────────
@app.get("/admin/cost-model")
async def list_cost_model(current_user: User = Depends(require_admin),
                          db: Session = Depends(get_db)):
    rows = db.query(CostModelEntry).order_by(CostModelEntry.pid).all()
    return [
        {
            "id": r.id, "branchName": r.branch_name, "glCode": r.gl_code,
            "branchCode": r.branch_code, "pid": r.pid, "glCategory": r.gl_category,
            "costModelCategory": r.cost_model_category, "description": r.description,
            "required": r.required, "currentCostModel": r.current_cost_model,
            "allocation": r.allocation, "futureCostModel": r.future_cost_model,
            "showbackType": r.showback_type, "userListingFlag": r.user_listing_flag,
        }
        for r in rows
    ]


@app.put("/admin/cost-model/{entry_id}")
async def update_cost_model(entry_id: int, body: dict,
                             current_user: User = Depends(require_admin),
                             db: Session = Depends(get_db)):
    entry = db.query(CostModelEntry).filter(CostModelEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(404, "Cost model entry not found")
    field_map = {
        "description": "description", "required": "required",
        "currentCostModel": "current_cost_model", "allocation": "allocation",
        "futureCostModel": "future_cost_model", "showbackType": "showback_type",
        "userListingFlag": "user_listing_flag", "costModelCategory": "cost_model_category",
    }
    for js_key, db_attr in field_map.items():
        if js_key in body:
            setattr(entry, db_attr, body[js_key])
    db.commit()
    # Re-run allocation so Cost Management tab reflects the updated Cost Model immediately.
    from allocator import run_allocation_all_periods
    rows = run_allocation_all_periods(db)
    if rows:
        existing = _load_data()
        _save_data(rows, existing.get("sheetName", "OC Data Refresh"))
    return {"ok": True}


# ── Admin: Headcount reference table ──────────────────────────────────────────
@app.get("/admin/headcount")
async def list_headcount(current_user: User = Depends(require_admin),
                         db: Session = Depends(get_db)):
    rows = db.query(HeadcountEntry).order_by(HeadcountEntry.short_code).all()
    return [{"id": r.id, "deptCode": r.dept_code, "shortCode": r.short_code,
             "deptName": r.dept_name,
             "fy2026": r.fy2026, "fy2027": r.fy2027, "fy2028": r.fy2028} for r in rows]


@app.post("/admin/headcount", status_code=201)
async def create_headcount(body: dict,
                            current_user: User = Depends(require_admin),
                            db: Session = Depends(get_db)):
    entry = HeadcountEntry(
        dept_code  = str(body.get("deptCode", "")),
        short_code = str(body.get("shortCode", "")),
        dept_name  = str(body.get("deptName", "")),
        fy2026     = float(body.get("fy2026", 0) or 0),
        fy2027     = float(body.get("fy2027", 0) or 0),
        fy2028     = float(body.get("fy2028", 0) or 0),
    )
    db.add(entry); db.commit(); db.refresh(entry)
    return {"ok": True, "id": entry.id}


@app.put("/admin/headcount/{entry_id}")
async def update_headcount(entry_id: int, body: dict,
                            current_user: User = Depends(require_admin),
                            db: Session = Depends(get_db)):
    entry = db.query(HeadcountEntry).filter(HeadcountEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(404, "Headcount entry not found")
    if "deptCode"  in body: entry.dept_code  = str(body["deptCode"])
    if "shortCode" in body: entry.short_code = str(body["shortCode"])
    if "deptName"  in body: entry.dept_name  = str(body["deptName"])
    for yr in ("fy2026", "fy2027", "fy2028"):
        if yr in body:
            try:
                setattr(entry, yr, float(body[yr]))
            except (ValueError, TypeError):
                raise HTTPException(400, f"{yr} must be a number")
    db.commit()
    from allocator import run_allocation_all_periods
    rows = run_allocation_all_periods(db)
    if rows:
        existing = _load_data()
        _save_data(rows, existing.get("sheetName", "OC Data Refresh"))
    return {"ok": True}


@app.delete("/admin/headcount/{entry_id}")
async def delete_headcount(entry_id: int,
                            current_user: User = Depends(require_admin),
                            db: Session = Depends(get_db)):
    entry = db.query(HeadcountEntry).filter(HeadcountEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(404, "Headcount entry not found")
    db.delete(entry); db.commit()
    return {"ok": True}


# ── Admin: User Listing reference table ───────────────────────────────────────
@app.get("/admin/user-listing")
async def list_user_listing(current_user: User = Depends(get_current_user),
                             db: Session = Depends(get_db)):
    if not current_user.is_admin and not current_user.can_edit_user_listing:
        raise HTTPException(403, "Admin or User Listing edit access required")
    query = db.query(UserListingEntry)
    if not current_user.is_admin and current_user.can_edit_user_listing:
        branches = [b.upper() for b in (current_user.allowed_branches or [])]
        if branches:
            query = query.filter(UserListingEntry.branch_code.in_(branches))
    rows = query.order_by(UserListingEntry.pid).all()
    return [
        {
            "id": r.id, "branchName": r.branch_name, "branchCode": r.branch_code,
            "glCode": r.gl_code, "pid": r.pid, "description": r.description,
            "ceo": r.ceo, "legal": r.legal, "corpOps": r.corp_ops, "hr": r.hr,
            "audit": r.audit, "cdo": r.cdo, "finance": r.finance,
            "technology": r.technology, "io": r.io, "irr": r.irr,
            "pe": r.pe, "cmci": r.cmci, "isr": r.isr,
        }
        for r in rows
    ]


@app.put("/admin/user-listing/{entry_id}")
async def update_user_listing(entry_id: int, body: dict,
                               current_user: User = Depends(get_current_user),
                               db: Session = Depends(get_db)):
    if not current_user.is_admin and not current_user.can_edit_user_listing:
        raise HTTPException(403, "Admin or User Listing edit access required")
    entry = db.query(UserListingEntry).filter(UserListingEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(404, "User listing entry not found")
    if not current_user.is_admin and current_user.can_edit_user_listing:
        branches = [b.upper() for b in (current_user.allowed_branches or [])]
        if branches and entry.branch_code.upper() not in branches:
            raise HTTPException(403, "You can only edit User Listing entries for your branches")
    field_map = {
        "ceo": "ceo", "legal": "legal", "corpOps": "corp_ops", "hr": "hr",
        "audit": "audit", "cdo": "cdo", "finance": "finance",
        "technology": "technology", "io": "io", "irr": "irr",
        "pe": "pe", "cmci": "cmci", "isr": "isr",
    }
    for js_key, db_attr in field_map.items():
        if js_key in body:
            try:
                setattr(entry, db_attr, float(body[js_key]))
            except (ValueError, TypeError):
                pass
    db.commit()
    from allocator import run_allocation_all_periods
    rows = run_allocation_all_periods(db)
    if rows:
        existing = _load_data()
        _save_data(rows, existing.get("sheetName", "OC Data Refresh"))
    return {"ok": True}


# ── Excel exports ────────────────────────────────────────────────────────────
@app.get("/admin/cost-model/export")
async def export_cost_model(current_user: User = Depends(require_admin),
                             db: Session = Depends(get_db)):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Model"
    headers = ['Branch Name', 'GL Code', 'Branch Code', 'PID', 'GL Category',
               'Cost Model Category', 'Description', 'Required', 'Current Cost Model',
               'Allocation', 'Future Cost Model', 'Showback Type', 'User Listing Flag']
    fill = PatternFill(start_color="00365B", end_color="00365B", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
    for e in db.query(CostModelEntry).all():
        ws.append([e.branch_name, e.gl_code, e.branch_code, e.pid, e.gl_category,
                   e.cost_model_category, e.description, e.required, e.current_cost_model,
                   e.allocation, e.future_cost_model, e.showback_type, e.user_listing_flag])
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(len(str(c.value or '')) for c in col) + 2, 50)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=cost_model.xlsx"})


@app.get("/admin/user-listing/export")
async def export_user_listing(current_user: User = Depends(require_admin),
                               db: Session = Depends(get_db)):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "User Listing"
    headers = ['Branch Name', 'Branch Code', 'GL Code', 'PID', 'Description',
               'CEO', 'Legal', 'Corp Ops', 'HR', 'Audit', 'CD&O',
               'Finance', 'Technology', 'IO', 'IRR', 'PE', 'CM&CI', 'ISR']
    fill = PatternFill(start_color="00365B", end_color="00365B", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
    for e in db.query(UserListingEntry).order_by(UserListingEntry.pid).all():
        ws.append([e.branch_name, e.branch_code, e.gl_code, e.pid, e.description,
                   e.ceo, e.legal, e.corp_ops, e.hr, e.audit, e.cdo,
                   e.finance, e.technology, e.io, e.irr, e.pe, e.cmci, e.isr])
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(len(str(c.value or '')) for c in col) + 2, 50)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=user_listing.xlsx"})


# ── Recalculate ───────────────────────────────────────────────────────────────
@app.post("/api/recalculate")
async def recalculate(request: Request, current_user: User = Depends(require_admin),
                      db: Session = Depends(get_db)):
    from allocator import run_allocation_all_periods
    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    base_year = int(body.get("baseYear", 2026))
    rows = run_allocation_all_periods(db, base_year=base_year)
    if not rows:
        raise HTTPException(400, "No OC data in database — upload a workbook first")
    payload = _load_data()
    _save_data(rows, payload.get("sheetName", "OC Data Refresh"))
    log_access(db, current_user.email, "recalculate", f"{len(rows)} rows",
               request.client.host if request.client else "")
    return {"ok": True, "rowCount": len(rows)}


# ── Share configs ────────────────────────────────────────────────────────────
_DEPT_LABELS = {
    'ceo': 'CEO', 'legal': 'Legal', 'hr': 'HR', 'audit': 'Audit',
    'cdo': 'CD&O', 'corpOps': 'Corp Ops', 'finance': 'Finance',
    'technology': 'Technology', 'io': 'IO', 'irr': 'IRR',
    'pe': 'PE', 'cmci': 'CM&CI', 'isr': 'ISR',
}


def _share_dict(c: ShareConfig) -> dict:
    return {
        "id":         c.id,
        "department": c.department,
        "emails":     c.emails or [],
        "schedule":   c.schedule,
        "period":     c.period,
        "lastSent":   c.last_sent.isoformat() if c.last_sent else None,
        "nextRun":    c.next_run.isoformat()  if c.next_run  else None,
    }


@app.get("/admin/share")
async def list_share_configs(current_user: User = Depends(require_admin),
                              db: Session = Depends(get_db)):
    return [_share_dict(c) for c in db.query(ShareConfig).order_by(ShareConfig.id).all()]


@app.post("/admin/share", status_code=201)
async def create_share_config(body: dict,
                               current_user: User = Depends(require_admin),
                               db: Session = Depends(get_db)):
    cfg = ShareConfig(
        department = str(body.get("department", "finance")),
        emails     = body.get("emails", []),
        schedule   = str(body.get("schedule", "manual")),
        period     = str(body.get("period", "actuals")),
    )
    db.add(cfg); db.commit(); db.refresh(cfg)
    return _share_dict(cfg)


@app.put("/admin/share/{cfg_id}")
async def update_share_config(cfg_id: int, body: dict,
                               current_user: User = Depends(require_admin),
                               db: Session = Depends(get_db)):
    cfg = db.query(ShareConfig).filter(ShareConfig.id == cfg_id).first()
    if not cfg:
        raise HTTPException(404, "Share config not found")
    if "department" in body: cfg.department = str(body["department"])
    if "emails"     in body: cfg.emails     = body["emails"]
    if "schedule"   in body: cfg.schedule   = str(body["schedule"])
    if "period"     in body: cfg.period     = str(body["period"])
    # Recompute next_run when schedule changes
    if "schedule" in body and body["schedule"] != "manual":
        cfg.next_run = _next_run(body["schedule"])
    elif "schedule" in body and body["schedule"] == "manual":
        cfg.next_run = None
    db.commit()
    return _share_dict(cfg)


@app.delete("/admin/share/{cfg_id}", status_code=204)
async def delete_share_config(cfg_id: int,
                               current_user: User = Depends(require_admin),
                               db: Session = Depends(get_db)):
    cfg = db.query(ShareConfig).filter(ShareConfig.id == cfg_id).first()
    if not cfg:
        raise HTTPException(404, "Share config not found")
    db.delete(cfg); db.commit()


@app.get("/admin/share/{cfg_id}/pdf")
async def preview_share_pdf(cfg_id: int,
                             current_user: User = Depends(require_admin),
                             db: Session = Depends(get_db)):
    from reporter import generate_dept_pdf
    from fastapi.responses import StreamingResponse
    import io
    cfg = db.query(ShareConfig).filter(ShareConfig.id == cfg_id).first()
    if not cfg:
        raise HTTPException(404, "Share config not found")
    data = _load_data()
    pdf  = generate_dept_pdf(data.get("rows", []), cfg.department, cfg.period)
    dept_label = _DEPT_LABELS.get(cfg.department, cfg.department.title())
    filename   = f"showback_{dept_label.lower().replace(' ', '_')}_{cfg.period}.pdf"
    return StreamingResponse(io.BytesIO(pdf),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"})


@app.post("/admin/share/{cfg_id}/send")
async def send_share_now(cfg_id: int, request: Request,
                          current_user: User = Depends(require_admin),
                          db: Session = Depends(get_db)):
    from reporter import generate_dept_pdf
    from emailer import send_report
    cfg = db.query(ShareConfig).filter(ShareConfig.id == cfg_id).first()
    if not cfg:
        raise HTTPException(404, "Share config not found")
    if not cfg.emails:
        raise HTTPException(400, "No email addresses configured")
    data = _load_data()
    if not data.get("rows"):
        raise HTTPException(400, "No cost data loaded — upload a workbook first")
    try:
        pdf = generate_dept_pdf(data["rows"], cfg.department, cfg.period)
        dept_label = _DEPT_LABELS.get(cfg.department, cfg.department.title())
        send_report(cfg.emails, dept_label, cfg.period, pdf)
    except RuntimeError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        raise HTTPException(500, f"Send failed: {e}")
    now = datetime.utcnow()
    cfg.last_sent = now
    if cfg.schedule != "manual":
        cfg.next_run = _next_run(cfg.schedule, now)
    db.commit()
    log_access(db, current_user.email, "share_send",
               f"{cfg.department} → {len(cfg.emails)} recipients",
               request.client.host if request.client else "")
    return {"ok": True, "recipients": len(cfg.emails)}


# ── Debug: show CM index entry for a given pid ────────────────────────────────
@app.get("/api/debug/cm/{pid}")
async def debug_cm(pid: str, current_user: User = Depends(require_admin),
                   db: Session = Depends(get_db)):
    from database import OcRawRow, CostModelEntry
    from allocator import parse_oc_cell, build_cost_model_index

    # Find OC raw rows that contain this pid
    pid_upper = pid.upper()
    oc_rows = db.query(OcRawRow).all()
    matched_oc = []
    for r in oc_rows:
        parsed = parse_oc_cell(r.oc_cell)
        if parsed and parsed.get('prod_id', '').upper() == pid_upper:
            matched_oc.append({
                'oc_cell': r.oc_cell,
                'parsed': parsed,
                'lookup_key': f"{parsed['level_code']}|{parsed['acct_num']}|{parsed['prod_id']}".lower(),
            })

    # Find CM entries with this pid (all of them)
    cm_entries = db.query(CostModelEntry).filter(
        CostModelEntry.pid.ilike(f'%{pid}%')
    ).all()
    cm_data = [{
        'id': e.id,
        'branch_name': e.branch_name,
        'branch_code': e.branch_code,
        'gl_code': e.gl_code,
        'pid': e.pid,
        'cm_key': f"{(e.branch_code or '').strip()}|{(e.gl_code or '').strip()}|{(e.pid or '').strip()}".lower(),
        'future_cost_model': e.future_cost_model,
        'current_cost_model': e.current_cost_model,
        'showback_type': e.showback_type,
    } for e in cm_entries]

    return {"oc_matches": matched_oc, "cm_entries": cm_data}


# ── Reset ─────────────────────────────────────────────────────────────────────
@app.post("/api/reset")
async def reset_all_data(request: Request, current_user: User = Depends(require_admin),
                         db: Session = Depends(get_db)):
    """Clear all cost data and reference tables — returns to a clean slate."""
    from database import OcRawRow
    db.query(OcRawRow).delete()
    db.query(CostModelEntry).delete()
    db.query(HeadcountEntry).delete()
    db.query(UserListingEntry).delete()
    db.commit()
    DATA_FILE.write_text('{"rows":[],"sheetName":null,"updatedAt":null}', encoding="utf-8")
    log_access(db, current_user.email, "reset_data", "all tables",
               request.client.host if request.client else "")
    return {"ok": True}


# ── Admin: app settings ───────────────────────────────────────────────────────
@app.get("/admin/settings")
async def get_settings(current_user: User = Depends(require_admin),
                       db: Session = Depends(get_db)):
    rows = db.query(AppSetting).all()
    return {r.key: r.value for r in rows}


@app.put("/admin/settings")
async def update_settings(body: dict,
                          current_user: User = Depends(require_admin),
                          db: Session = Depends(get_db)):
    for key, value in body.items():
        setting = db.query(AppSetting).filter(AppSetting.key == key).first()
        if setting:
            setting.value = str(value)
        else:
            db.add(AppSetting(key=key, value=str(value)))
    db.commit()
    return {"ok": True}


# ── Health ────────────────────────────────────────────────────────────────────
@app.get("/health")
async def health():
    return {"ok": True, "version": "1.0.0"}


# ── Frontend static serving (must be last — catch-all would shadow API routes) ─
if DIST_DIR.exists():
    app.mount("/assets", StaticFiles(directory=str(DIST_DIR / "assets")), name="static-assets")

@app.get("/{full_path:path}")
async def spa_fallback(full_path: str):
    index = DIST_DIR / "index.html"
    if index.exists():
        return FileResponse(str(index))
    return JSONResponse({"error": "frontend not built — run: npm run build"}, status_code=404)
